"""
XLSX export -- a proper multi-sheet workbook rather than a single flat
table, per spec section 42: Summary, Students (results), Answers,
Question Analysis.
"""
from __future__ import annotations

import io
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.analytics.engine import compute_exam_stats, compute_question_stats
from app.core.models import Exam
from app.database import repository as repo

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _autosize(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(40, max(10, max_len + 2))


def export_exam_workbook(conn: sqlite3.Connection, exam: Exam) -> bytes:
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    stats = compute_exam_stats(conn, exam.id)
    rows = [
        ("Exam", exam.name),
        ("Subject", exam.subject),
        ("Question count", exam.question_count),
        ("Option count", exam.option_count),
        ("Scoring: correct", exam.scoring.correct),
        ("Scoring: wrong", exam.scoring.wrong),
        ("Scoring: blank", exam.scoring.blank),
        ("Multiple-mark policy", exam.scoring.multiple_mark_policy),
        ("", ""),
        ("Submissions", stats.n_submissions),
        ("Scored", stats.n_scored),
        ("Average %", stats.average),
        ("Median %", stats.median),
        ("Min %", stats.minimum),
        ("Max %", stats.maximum),
        ("Std dev", stats.std_dev),
    ]
    for r, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30

    # ---- Students (results) ----
    ws2 = wb.create_sheet("Students")
    headers = ["Submission ID", "Student ID", "Name", "File", "Status", "Quality %", "Score", "Percentage", "Timestamp"]
    _write_header(ws2, 1, headers)
    submissions = repo.list_submissions(conn, exam.id)
    for r, s in enumerate(submissions, start=2):
        ws2.cell(row=r, column=1, value=s.id)
        ws2.cell(row=r, column=2, value=s.student_id_effective or "")
        ws2.cell(row=r, column=3, value=s.student_name_effective or "")
        ws2.cell(row=r, column=4, value=s.source_file)
        ws2.cell(row=r, column=5, value=s.status.value)
        ws2.cell(row=r, column=6, value=s.quality_score)
        ws2.cell(row=r, column=7, value=s.score)
        ws2.cell(row=r, column=8, value=s.percentage)
        ws2.cell(row=r, column=9, value=str(s.timestamp) if s.timestamp else "")
    _autosize(ws2, len(headers))

    # ---- Answers ----
    ws3 = wb.create_sheet("Answers")
    headers3 = ["Submission ID", "Question", "Detected", "Final (if corrected)", "Correct Answer", "Confidence", "Status"]
    _write_header(ws3, 1, headers3)
    answer_rows = repo.get_all_answer_results_for_exam(conn, exam.id)
    for r, row in enumerate(answer_rows, start=2):
        ws3.cell(row=r, column=1, value=row["submission_id"])
        ws3.cell(row=r, column=2, value=row["question_number"])
        ws3.cell(row=r, column=3, value=row["detected_answer"] or "")
        ws3.cell(row=r, column=4, value=row["final_answer"] or "")
        ws3.cell(row=r, column=5, value=row["correct_answer"] or "")
        ws3.cell(row=r, column=6, value=row["confidence"])
        ws3.cell(row=r, column=7, value=row["status"])
    _autosize(ws3, len(headers3))

    # ---- Question Analysis ----
    ws4 = wb.create_sheet("Question Analysis")
    headers4 = ["Question", "Correct Answer", "Correct %", "Wrong %", "Blank %", "Multiple %", "Review Rate %", "Avg Confidence", "N"]
    _write_header(ws4, 1, headers4)
    for r, qs in enumerate(compute_question_stats(conn, exam.id), start=2):
        ws4.cell(row=r, column=1, value=qs.question_number)
        ws4.cell(row=r, column=2, value=qs.correct_answer or "")
        ws4.cell(row=r, column=3, value=qs.correct_pct)
        ws4.cell(row=r, column=4, value=qs.wrong_pct)
        ws4.cell(row=r, column=5, value=qs.blank_pct)
        ws4.cell(row=r, column=6, value=qs.multiple_pct)
        ws4.cell(row=r, column=7, value=qs.review_rate_pct)
        ws4.cell(row=r, column=8, value=qs.avg_confidence)
        ws4.cell(row=r, column=9, value=qs.n)
    _autosize(ws4, len(headers4))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
