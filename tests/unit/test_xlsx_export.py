import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.database.db import connect, init_db
from app.database import repository as repo
from app.core.models import AnswerResult, AnswerStatus, Exam, ScoringRule, Submission, SubmissionStatus
from app.export.xlsx_export import export_exam_workbook


class TestXlsxExport(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.TemporaryDirectory()
        self.db_path = Path(self._tmpdir.name) / "test.sqlite3"
        self.conn = connect(self.db_path)
        init_db(self.conn)
        self.exam = Exam(name="XLSX Test Exam", subject="Chem", question_count=2, option_count=4,
                          scoring=ScoringRule(correct=1, wrong=-0.25))
        self.exam.id = repo.create_exam(self.conn, self.exam)
        repo.set_answer_key(self.conn, self.exam.id, {1: "A", 2: "B"})
        sub_id = repo.create_submission(self.conn, Submission(
            exam_id=self.exam.id, source_file="s1.png", student_id_detected="1001",
            status=SubmissionStatus.COMPLETED, score=0.75, percentage=75.0))
        repo.save_answer_results(self.conn, sub_id, [
            AnswerResult(sub_id, 1, "A", 92.0, AnswerStatus.HIGH_CONFIDENCE),
            AnswerResult(sub_id, 2, "C", 88.0, AnswerStatus.HIGH_CONFIDENCE),
        ])

    def tearDown(self):
        self.conn.close()
        self._tmpdir.cleanup()

    def test_workbook_has_all_required_sheets(self):
        data = export_exam_workbook(self.conn, self.exam)
        self.assertGreater(len(data), 0)
        wb = load_workbook(io_bytes(data))
        self.assertEqual(set(wb.sheetnames), {"Summary", "Students", "Answers", "Question Analysis"})

    def test_summary_sheet_has_real_exam_data(self):
        data = export_exam_workbook(self.conn, self.exam)
        wb = load_workbook(io_bytes(data))
        ws = wb["Summary"]
        values = {row[0].value: row[1].value for row in ws.iter_rows() if row[0].value}
        self.assertEqual(values["Exam"], "XLSX Test Exam")
        self.assertEqual(values["Submissions"], 1)

    def test_students_sheet_has_the_submission(self):
        data = export_exam_workbook(self.conn, self.exam)
        wb = load_workbook(io_bytes(data))
        ws = wb["Students"]
        header = [c.value for c in ws[1]]
        self.assertIn("Student ID", header)
        row2 = [c.value for c in ws[2]]
        self.assertIn("1001", row2)

    def test_answers_sheet_reflects_key_and_detection(self):
        data = export_exam_workbook(self.conn, self.exam)
        wb = load_workbook(io_bytes(data))
        ws = wb["Answers"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        q2 = [r for r in rows if r[1] == 2][0]
        self.assertEqual(q2[2], "C")   # detected
        self.assertEqual(q2[4], "B")   # correct answer from key

    def test_question_analysis_sheet_has_both_questions(self):
        data = export_exam_workbook(self.conn, self.exam)
        wb = load_workbook(io_bytes(data))
        ws = wb["Question Analysis"]
        q_numbers = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
        self.assertEqual(q_numbers, [1, 2])


def io_bytes(data: bytes):
    import io
    return io.BytesIO(data)


if __name__ == "__main__":
    unittest.main()
